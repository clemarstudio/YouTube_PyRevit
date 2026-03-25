# --- [1] IMPORTS ---
# We use xml.etree.ElementTree to read and modify Navisworks XML files
import xml.etree.ElementTree as ET
# openpyxl is used to create and read Excel (.xlsx) files
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os
import sys
# tkinter is used for simple pop-up windows to pick files
from tkinter import filedialog, Tk, messagebox

"""`
NAVISWORKS CLASH NAME EDITOR (Tutorial Version)
Workflow:
1. Export: Select an .xml export from Navisworks -> This script creates an .xlsx file.
2. Edit: Open the Excel, change names in the "New Name" column, and save.
3. Import: Select the updated .xlsx file -> This script creates a new "_Renamed.xml".
"""

# Registering namespace ensures the XML output format matches what Navisworks expects
ET.register_namespace("xsi", "http://www.w3.org/2001/XMLSchema-instance")

# --- [2] CORE CLASS (THE ENGINE) ---
class ClashRenamer:
    def __init__(self, xml_path):
        self.xml_path = xml_path
        self.tree = None
        self.root = None

    def load_xml(self):
        """Loads and parses the Navisworks XML file into a Python object."""
        if not os.path.exists(self.xml_path):
            return False
        try:
            # ET.parse turns the raw file into a 'Tree' we can navigate
            self.tree = ET.parse(self.xml_path)
            self.root = self.tree.getroot()
            return True
        except Exception as e:
            print(f"Error parsing XML: {e}")
            return False

    def export_to_excel(self, excel_path):
        """Finds all Clash Tests in the XML and writes them to Excel."""
        if self.root is None and not self.load_xml():
            return

        # Initialize a new Excel Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clash Names"

        # Headers help the user know where to type
        headers = ["ID (Do Not Edit)", "Original Name", "New Name (Edit This)"]
        ws.append(headers)

        # Apply basic styling to headers (Blue background, White text)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Iterate through the XML looking for 'clashtest' elements
        row_idx = 2
        for i, clashtest in enumerate(self.root.iter('clashtest')):
            # Grab the 'name' attribute from the XML tag
            original_name = clashtest.get('name', '')
            
            # Write to Excel columns
            ws.cell(row=row_idx, column=1, value=i)              # A: Index
            ws.cell(row=row_idx, column=2, value=original_name)  # B: Current Name
            ws.cell(row=row_idx, column=3, value=original_name)  # C: Future Name
            row_idx += 1

        # Adjust widths so the names are easier to read
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        # Hide Column A because users shouldn't change the Index IDs
        ws.column_dimensions['A'].hidden = True 

        wb.save(excel_path)
        print(f"✅ Exported {row_idx - 2} tests to {excel_path}")

    def import_from_excel(self, excel_path, output_xml_path):
        """Reads the edited Excel and updates the XML 'name' attributes."""
        if self.root is None and not self.load_xml():
            return

        # Load the Excel file specifically to read values
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # We store the user's new names in a Dictionary (Index -> New Name)
        rename_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            idx, old_name, new_name = row
            if idx is not None and new_name:
                rename_map[int(idx)] = str(new_name)

        # Loop through the XML again and replace names based on the Index
        updated_count = 0
        for i, clashtest in enumerate(self.root.iter('clashtest')):
            if i in rename_map:
                clashtest.set('name', rename_map[i])
                updated_count += 1

        # Save the modified tree to a new XML file
        self.tree.write(output_xml_path, encoding="utf-8", xml_declaration=True)
        print(f"✅ Updated {updated_count} names! Created: {output_xml_path}")

# --- [3] APPLICATION LOGIC (THE CONTROLLER) ---
def main():
    # Hide the main Tkinter window (we only want the file dialog)
    root = Tk()
    root.withdraw()

    # User picks a file via a Windows explorer pop-up
    file_path = filedialog.askopenfilename(
        title="Select Navisworks XML (to Export) or Excel (to Import)",
        filetypes=[("Navisworks/Excel", "*.xml *.xlsx")]
    )

    if not file_path:
        return

    # Check the file extension to decide what to do
    ext = os.path.splitext(file_path)[1].lower()
    base_name = os.path.splitext(file_path)[0]

    try:
        if ext == ".xml":
            # SCENARIO A: XML -> EXCEL
            excel_out = base_name + ".xlsx"
            renamer = ClashRenamer(file_path)
            renamer.export_to_excel(excel_out)
            messagebox.showinfo("Success", f"Excel Exported to:\n{excel_out}")

        elif ext == ".xlsx":
            # SCENARIO B: EXCEL -> XML
            # Assumption: The original XML has the same name as the Excel
            xml_original_source = base_name + ".xml"
            xml_final_output = base_name + "_Renamed.xml"
            
            if not os.path.exists(xml_original_source):
                messagebox.showerror("Error", f"XML source not found:\n{xml_original_source}")
                return

            renamer = ClashRenamer(xml_original_source)
            renamer.import_from_excel(file_path, xml_final_output)
            messagebox.showinfo("Success", f"New XML Created:\n{xml_final_output}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{e}")

# Start the script
if __name__ == "__main__":
    main()
